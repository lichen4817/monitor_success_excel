#!/usr/bin/python
# -*- coding: UTF-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment, Border, Side
import numpy as np
import xlrd as xd
import xlwt as xw
from win32com.client import constants as c
import string
import sys
inputname=sys.argv[1]
inputdir=sys.argv[2]
outputdir=sys.argv[3]



date_file = inputname[len(inputname) - 9:len(inputname) - 5]
# 确定结果sheet名
result_sheetname = 'LCD点位推送及回传数据分析' + date_file
impt_xlsx_resource = pd.read_excel(inputdir + inputname, sheet_name='销售城市')
# impt_xlsx_resource.replace('不分套装', '合计', inplace=True)
sheet4_dataframe = impt_xlsx_resource[(impt_xlsx_resource.媒体 == 'lcd') & (impt_xlsx_resource.套装 == '不分套装') & (impt_xlsx_resource.城市 != '无匹配')].pivot_table(
    index='城市',
    values=['安装总数',
            '自有智能屏在线总数',
            '自有智能屏不在线总数',
            '自有非智能屏幕13寸总数',
            '自有非智能屏幕非13寸总数',
            '租用总数',
            '推送成功总数',
            '全设备推送成功率',
            '智能屏推送成功率',
            '小时统计推送成功且七天平均>20h比率'
            ])
sheet4_dataframe = sheet4_dataframe.sort_values('安装总数', ascending=False)

# 重命名字段名
sheet4_dataframe.rename(columns={'小时统计推送成功且七天平均>20h比率': '可在线监测比'}, inplace=True)

# 重置列的顺序
order = ['安装总数',
         '自有智能屏在线总数',
         '自有智能屏不在线总数',
         '自有非智能屏幕13寸总数',
         '自有非智能屏幕非13寸总数',
         '租用总数',
         '推送成功总数',
         '全设备推送成功率',
         '智能屏推送成功率',
         '可在线监测比']
sheet4_dataframe = sheet4_dataframe[order]
sheet4_dataframe['全设备推送成功率'] = sheet4_dataframe['全设备推送成功率'].apply(lambda x: format(x, '.2%'))
sheet4_dataframe['智能屏推送成功率'] = sheet4_dataframe['智能屏推送成功率'].apply(lambda x: format(x, '.2%'))
sheet4_dataframe['可在线监测比'] = sheet4_dataframe['可在线监测比'].apply(lambda x: format(x, '.2%'))


##返回行数
row_1 = sheet4_dataframe.shape[0]
##返回列数
column_1 = sheet4_dataframe.shape[1]
writer=pd.ExcelWriter(outputdir + 'monitor_reuslt' + date_file +'.xlsx',engine='openpyxl')
book = load_workbook(writer.path)
writer.book = book
sheet4_dataframe.to_excel(excel_writer=writer,sheet_name=result_sheetname)
writer.save()
writer.close()

##修改样式,遍历A-Z列
workbook = load_workbook(outputdir + 'monitor_reuslt' + date_file +'.xlsx')
ali = Alignment(horizontal='centerContinuous', vertical='center')
font = Font(name='微软雅黑', size=12, bold=False, italic=False)
border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                right=Side(border_style='thin'))
workbook._active_sheet_index=2
sheet = workbook.active
first_change = 1
for column_no in string.ascii_uppercase:
    for row_no in range(1, row_1 + 2):
        c = str(column_no) + str(row_no)
        cell = sheet[c]
        cell.font = font
        cell.alignment = ali
        cell.border = border
    first_change += 1
    if first_change > column_1 + 1:
        break
#改变格式后的文件反写会对应sheet
workbook.save(filename=outputdir + 'monitor_reuslt' + date_file +'.xlsx')
print(result_sheetname + ' Success !')

